from openpyxl import Workbook, load_workbook
import smtplib
import email
#import imaplib
from datetime import datetime
import re

def validateEmail(email):
    if '@' not in email:
        print('Enter a valid email address')
        e = input('Email: ')
        return validateEmail(e)
    else:
        e = email
        return e


def validateUse(use):
    use = use.lower()
    valid_use_inputs = ['use', 'u']
    valid_preview_inputs  = ['preview', 'p']
    if use not in valid_use_inputs and use not in valid_preview_inputs:
        print('Select a valid input: use, u, preview, or p')
        u = input('PREVIEW or Use: ')
        return validateUse(u)
    elif use in valid_use_inputs:
        return str('Use')
    elif use in valid_preview_inputs:
        return str('PREVIEW')


def getValues():
    first = input('Firstname: ')
    last = input('Lastname: ')
    city = input('City: ')

    email = input('Email: ')
    valid_email = validateEmail(email)

    use = input('PREVIEW or Use: ')
    valid_use = validateUse(use)

    username = first.lower()[0] + last.lower()

    values = {'first': first, 'last': last, 'city': city, 'email': valid_email, 'use': valid_use, 'username': username}

    return values


def updateUserList(user_workbook, values):
    user_information = load_workbook(user_workbook)
    ws = user_information.active

    count = 2825
    for col in ws.iter_cols(min_row=2825, max_col=1, max_row=5000):
        for cell in col:
            if cell.value is None:
                #print(cell)
                #print(count)
                last_pass_cell = ws['B{}'.format(count-1)]
                last_pass = last_pass_cell.value
                r = re.search('mssci(\d+)', str(last_pass))
                num = r.group(1)
                opn = int(num)
                #print(num)
                npn = opn + 1
                #print(npn)
                password = 'mssci{}'.format(npn) 

                ws['A{}'.format(count)] = values['username']
                ws['B{}'.format(count)] = password
                ws['C{}'.format(count)] = values['first']
                ws['D{}'.format(count)] = values['last']
                ws['E{}'.format(count)] = values['email']
                ws['H{}'.format(count)] = values['city']
                ws['I{}'.format(count)] = values['use']
                break
            count += 1

    #rng = 'A{}:I{}'.format(count)
    #new_entry_cells = ws[rng]
    #new_entry_cells.value = 'TEST'

    user_information.save(user_workbook)
    return password


def previewExcel(preview_workbook, values, password, pcount, pclear):
    pwb = load_workbook(preview_workbook)
    ws = pwb.active

    if pclear == False:
        for col in ws.iter_cols(min_row=2, max_col=6, max_row=10):
            for cell in col:
               cell.value = None 

    ws['A{}'.format(pcount)] = values['username']
    ws['B{}'.format(pcount)] = password
    ws['C{}'.format(pcount)] = values['first']
    ws['D{}'.format(pcount)] = values['last']
    ws['E{}'.format(pcount)] = values['email']

    pwb.save(preview_workbook)


def useExcel(use_workbook, values, password, ucount, uclear):
    pwb = load_workbook(use_workbook)
    ws = pwb.active

    if uclear == False:
        for col in ws.iter_cols(min_row=2, max_col=6, max_row=10):
            for cell in col:
               cell.value = None 

    ws['A{}'.format(ucount)] = values['username']
    ws['B{}'.format(ucount)] = password
    ws['C{}'.format(ucount)] = values['first']
    ws['D{}'.format(ucount)] = values['last']
    ws['E{}'.format(ucount)] = values['email']
    ws['F{}'.format(ucount)] = 'student{}'.format(values['last'].lower())

    pwb.save(use_workbook)


def addUserToMailMergeExcel(values):
    return values


def addAnotherUser(pcount, pclear, ucount, uclear):
    con = input('Add another user? [Y/n]')
    while con == 'Y' or con =='':
        values = getValues()
        password = updateUserList('/Users/cyee/Desktop/mssci USER LIST UPDATED.xlsx', values)

        if values['use'] == 'PREVIEW':
            previewExcel('/Users/cyee/Desktop/MSSCI MAIL MERGE PREVIEW.xlsx', values, password, pcount, pclear)
            pcount += 1
            pclear = True
        elif values['use'] == 'Use':
            useExcel('/Users/cyee/Desktop/MSSCI MAIL MERGE USER.xlsx', values, password, ucount, uclear)
            ucount += 1
            uclear = True

        values['password'] = password
        saveUsersForEmail(values, 'a')
        con = addAnotherUser(pcount, pclear, ucount, uclear)


def saveUsersForEmail(values, w_or_a):
    with open('./values.txt', w_or_a) as f:
       f.write(str(values) + '\n') 
       f.close()


if __name__ == '__main__':
    #read_mail()
    values = getValues()
    password = updateUserList('/Users/cyee/Desktop/mssci USER LIST UPDATED.xlsx', values)
    pcount = 2
    pclear = False
    ucount = 2
    uclear = False
    use_inputs = ['use', 'u']
    preview_inputs = ['preview', 'p']

    if values['use'] ==  'PREVIEW':
        previewExcel('/Users/cyee/Desktop/MSSCI MAIL MERGE PREVIEW.xlsx', values, password, pcount, pclear)
        pcount += 1
        pclear = True
    elif values['use'] == 'Use':
        useExcel('/Users/cyee/Desktop/MSSCI MAIL MERGE USER.xlsx', values, password, ucount, uclear)
        ucount += 1
        uclear = True

    values['password'] = password
    saveUsersForEmail(values, 'w')
    addAnotherUser(pcount, pclear, ucount, uclear)


